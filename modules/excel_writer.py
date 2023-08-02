from hashlib import sha1
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Color, Alignment, Border, Side
from modules.decorators import file_decorator


def generate_name():
    data = sha1(str(datetime.now()).encode())
    return str(data.hexdigest())


def set_main_styles(sheet:Worksheet):
    header_font = Font(bold=True, size=14)
    center_aligned_text = Alignment(horizontal="center")
    border_side = Side(border_style="thick")
    thick_border = Border(top=border_side, left=border_side, right=border_side, bottom=border_side)

    sheet['A1'].font = header_font
    sheet['A1'].alignment = center_aligned_text
    sheet['A1'].border = thick_border
    sheet['B1'].font = header_font
    sheet['B1'].alignment = center_aligned_text
    sheet['B1'].border = thick_border
    sheet['C1'].font = header_font
    sheet['C1'].alignment = center_aligned_text
    sheet['C1'].border = thick_border

    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 50


def set_cell_style(cell:Worksheet.cell):
    center_aligned_text = Alignment(horizontal="center")
    border_side = Side(border_style="thin")
    normal_font = Font(size=12)
    normal_border = Border(top=border_side, left=border_side, right=border_side, bottom=border_side)
    cell.alignment = center_aligned_text
    cell.border = normal_border
    cell.font = normal_font


@file_decorator('cant create file')
def wrtie_xls(data: dict, name: str):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Provider Output"
    sheet['A1'].value = "№ п/п"
    sheet['B1'].value = "Адрес места хранения"
    sheet['C1'].value = "Штрихкод"
    set_main_styles(sheet)
    count = 0
    for key in data.keys():
        count += 1
        c = sheet.cell(row=1+count, column=1)
        c.value = count
        set_cell_style(c)
        c = sheet.cell(row=1+count, column=2)
        c.value = key
        set_cell_style(c)
        c = sheet.cell(row=1+count, column=3)
        c.value = data[key]
        set_cell_style(c)
    wb.save(f"{name}.xlsx")
    return True
